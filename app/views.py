from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect

from .models import Questao, Prova, Candidato, Resposta

from random import shuffle

# Pegando GSPREAD CLIENTE pelo arquivo Settings
from sistema import settings
import gspread
import pandas as pd
import openpyxl


@login_required
def menu(request):
    return render(request, 'menu/index.html')


@login_required
def index_questao(request):
    questoes = Questao.objects.all()
    context = {
        'questoes' : questoes
    }
    return render(request, 'questao/index.html', context)


@login_required
def criar_questao(request):
    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        peso = request.POST.get('peso')
        dificuldade = request.POST.get('dificuldade')
        materia = request.POST.get('materia')
        texto_questao = request.POST.get('texto_questao')
        opcao_a = request.POST.get('opcao_a')
        opcao_b = request.POST.get('opcao_b')
        opcao_c = request.POST.get('opcao_c')
        opcao_d = request.POST.get('opcao_d')
        opcao_e = request.POST.get('opcao_e')
        correta = request.POST.get('correta')
        Questao.objects.create(titulo=titulo, peso=peso, dificuldade=dificuldade, materia=materia, texto_questao=texto_questao, opcao_a=opcao_a, opcao_b=opcao_b, opcao_c=opcao_c, opcao_d=opcao_d, opcao_e=opcao_e, correta=correta)
        return redirect(add_questao)
    else:
        return render(request, 'questao/criar.html')
    
    
def add_questao(request):
    return render(request, 'questao/adicionado.html')
    

@login_required
def editar_questao_get(request, id):
    questao = Questao.objects.get(id=id)
    
    context = {
        'questao' : questao
    }
    return render(request, 'questao/editar.html', context)


def editar_questao_post(request, id):
    questao = Questao.objects.get(id=id)
    questao.titulo = request.POST.get('titulo')
    questao.peso = request.POST.get('peso')
    questao.dificuldade = request.POST.get('dificuldade')
    questao.materia = request.POST.get('materia')
    questao.texto_questao = request.POST.get('texto_questao')
    questao.opcao_a = request.POST.get('opcao_a')
    questao.opcao_b = request.POST.get('opcao_b')
    questao.opcao_c = request.POST.get('opcao_c')
    questao.opcao_d = request.POST.get('opcao_d')
    questao.opcao_e = request.POST.get('opcao_e')
    questao.correta = request.POST.get('correta')
    questao.save()
    return redirect(index_questao)


def confirmar_deletar_questao(request, id):
    questao = Questao.objects.get(id=id)
    context = {
        'questao' : questao
    }
    return render(request, 'questao/deletar.html', context)


@login_required
def deletar_questao(request, id):
    questao = Questao.objects.get(id=id)
    questao.delete()
    return redirect(index_questao)


@login_required
def index_prova(request):
    provas = Prova.objects.all()
    context = {
        'provas' : provas
    }
    return render(request, 'prova/index.html', context)


@login_required
def criar_prova(request):
    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        curso = request.POST.get('curso')
        qnt_questoes = int(request.POST.get('quantidade'))
        questoes_total = list(Questao.objects.all())
        shuffle(questoes_total)
        questoes_sorteadas = questoes_total[:qnt_questoes]
        prova = Prova.objects.create(titulo=titulo, curso=curso, quantidade_questoes=qnt_questoes) 
        prova.questoes.set(questoes_sorteadas)
        return redirect(add_prova)
    else:
        return render(request, 'prova/criar.html')


@login_required
def add_prova(request):
    return render(request, 'prova/adicionado.html')


@login_required
def editar_prova(request, id):
    prova = Prova.objects.get(id=id)
    if request.method == 'POST':
        status = request.POST.get('status')
        prova.status = status
        prova.save()
        return redirect(index_prova)
    else:
        context = {
            'prova' : prova
        }
        return render(request, 'prova/editar.html', context)


@login_required
def confirmar_deletar_prova(request, id):
    prova = Prova.objects.get(id=id)
    context = {
        'prova' : prova
    }
    return render(request, 'prova/deletar.html', context)

@login_required
def deletar_prova(request, id):
    prova = Prova.objects.get(id=id)
    prova.delete()
    return redirect(index_prova)


@login_required
def index_dashboard(request):
    return render(request, 'dashboard/index.html')


def index(request):
    return render(request, 'candidato/index.html')


def capturar_informacoes(request):
        nome = request.POST.get('nome')
        email = request.POST.get('email')
        telefone = request.POST.get('telefone')
        if not Candidato.objects.filter(email=email).exists(): 
            Candidato.objects.create(nome=nome, email=email, telefone=telefone) 
        provas = Prova.objects.all()
        provas_s = list()
        for prova in provas:
            if prova.status:
                provas_s.append(prova)
        context = {
            'provas' : provas_s
        }
        response = render(request, 'candidato/provas.html', context)
        response.set_cookie('email', email)
        return response
    

def candidato_prova(request, id):
    candidato_email = request.COOKIES.get('email')
    candidato = Candidato.objects.get(email=candidato_email)
    prova = Prova.objects.get(id=id) 
    if request.method == 'POST':
        respostas_lista = []
        for questao in prova.questoes.all():
            nome_resposta = f'resposta_{questao.id}'
            nome_questao = f'questao_{questao.id}'
            peso_questao = f'{questao.peso}'
            correta_questao = f'{questao.correta}'
            resposta_candidato = request.POST.get(nome_resposta)            
            respostas_lista.append(nome_questao+ ',' + peso_questao + ',' + resposta_candidato + ',' + correta_questao)
        respostas = ','.join(respostas_lista)
        respostas = respostas.split(",")
        Resposta.objects.create(candidato=candidato, prova=prova, respostas=respostas)


        nome_planilha = "Avaliações BOSCH UVA"
        possivel_nome_folha = str(prova.titulo) + "-" + str(prova.id)

        # Abra a planilha pelo nome
        planilha = settings.GSPREAD_CLIENT.open(nome_planilha)

        # Verifique se a folha já existe
        while True:
            try:
                folha = planilha.worksheet(possivel_nome_folha)
                info_candidato = [candidato.nome, candidato.email, candidato.telefone, prova.titulo, prova.curso]
                for questao in range(prova.quantidade_questoes*4):
                    info_candidato.append(respostas[questao])

                folha.append_row(info_candidato)

                acertos = 0
                erros = 0
                pontuacao = 0
                linha_atual = len(folha.get_all_values())
                coluna_atual = (prova.quantidade_questoes*4)+5
                for coluna in range(8, coluna_atual, 4):
                    value_quest = folha.cell(row=linha_atual, col=coluna).value
                    value_resp = folha.cell(row=linha_atual, col=coluna + 1).value
                    peso = folha.cell(row=linha_atual, col=coluna-1).value

                    # Comparar e contar acertos
                    if value_quest == value_resp:
                        acertos += 1
                        pontuacao += 1 * int(peso)
                    else:
                        erros += 1

                # Registrar a contagem de acertos na nova coluna
                folha.update_cell(row=linha_atual, col=coluna_atual + 1, value=acertos)
                folha.update_cell(row=linha_atual, col=coluna_atual + 2, value=erros)
                folha.update_cell(row=linha_atual, col=coluna_atual + 3, value=pontuacao)

                df = pd.DataFrame(folha.get_all_values())
                nome_arquivo = possivel_nome_folha + ".xlsx"
                with pd.ExcelWriter("S:/PM/ter/ets/Inter_Setor/COMPARTILHADO/APRENDIZES/DIGITAL_SOLUTIONS_09/GABRIEL OLIVEIRA JORDAO/hackathon/"+nome_arquivo, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name=nome_arquivo, index=False, header=False, columns=None)
                break
            except:
                folha = planilha.add_worksheet(title=possivel_nome_folha, rows=100, cols=100)

                cabecalho = ["Nome", "Email", "Telefone", "Título da Pova", "Curso"]
                for id in range(prova.quantidade_questoes):
                    cabecalho.append("Questão " + str(id+1))
                    cabecalho.append("Peso")
                    cabecalho.append("Resposta do Candidato")
                    cabecalho.append("Resposta Correta")
                
                cabecalho.append("Acertos")
                cabecalho.append("Erros")
                cabecalho.append("Pontuação")
                folha.insert_row(cabecalho, index=1)
                
        return render(request, 'candidato/enviado.html')
    else:
        questoes = prova.questoes.all()
        context = {
            'prova': prova,
            'questoes': questoes,
            'candidato': candidato
        }
        return render(request, 'candidato/prova.html', context)


"""Checar se está sendo usada"""
def candidato_finalizada(request):
    return render(request, 'candidato/enviado.html')
    
